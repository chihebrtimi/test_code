# monitoring/admin.py
from django.contrib import admin
from django.http import HttpResponse
from django.contrib.auth.admin import UserAdmin
from django.contrib.auth.models import Group
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from django.utils.timezone import now
from datetime import datetime, timedelta

from .models import SensorData, CustomUser, LoginHistory, Message


# 🔴 Custom AdminSite
class CustomAdminSite(admin.AdminSite):
    site_header = "Admin Dashboard"
    site_title = "Monitoring Admin"

    def each_context(self, request):
        context = super().each_context(request)
        context["custom_admin_css"] = "/static/css/custom_admin.css"
        return context

    def get_app_list(self, request, app_label=None):
        app_list = super().get_app_list(request, app_label)
        for app in app_list:
            if app["app_label"] == "monitoring":
                app["name"] = "Monitoring Administration"
                for model in app["models"]:
                    model["app_label"] = ""
        return app_list


# Replace default admin site with ours
custom_admin_site = CustomAdminSite(name="custom_admin")

# 🚫 Unregister Groups (not needed anymore)
try:
    custom_admin_site.unregister(Group)
except admin.sites.NotRegistered:
    pass


# ✅ Export selected SensorData to Excel
def export_to_excel(modeladmin, request, queryset):
    wb = Workbook()
    ws = wb.active
    ws.title = "Sensor Data"

    field_names = [
        'timestamp',
        'T_cold', 'Humidity_cold',
        'T_hot_rack1', 'Humidity_rack1',
        'T_hot_rack2', 'Humidity_rack2',
        'T_hot_rack3', 'Humidity_rack3',
        'T_room', 'Room_Humidity',
        'P_total_room', 'P_total_cooling_system',
        'P_rack1', 'P_rack2'
    ]

    for col_num, field in enumerate(field_names, 1):
        ws[f'{get_column_letter(col_num)}1'] = field

    for row_num, obj in enumerate(queryset, 2):
        for col_num, field in enumerate(field_names, 1):
            ws.cell(row=row_num, column=col_num, value=getattr(obj, field))

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=SensorDataExport.xlsx'
    wb.save(response)
    return response

export_to_excel.short_description = "📤 Export selected to Excel"

# ✅ Export Messages to Excel
def export_messages_to_excel(modeladmin, request, queryset):
    wb = Workbook()
    ws = wb.active
    ws.title = "Messages"

    field_names = ['timestamp', 'type', 'content', 'data', 'source']

    for col_num, field in enumerate(field_names, 1):
        ws[f'{get_column_letter(col_num)}1'] = field

    for row_num, obj in enumerate(queryset, 2):
        ws.cell(row=row_num, column=1, value=obj.timestamp.strftime("%Y-%m-%d %H:%M:%S"))
        ws.cell(row=row_num, column=2, value=obj.type)
        ws.cell(row=row_num, column=3, value=obj.content)
        ws.cell(row=row_num, column=4, value=str(obj.data) if obj.data else "")
        ws.cell(row=row_num, column=5, value=obj.source)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=MessagesExport.xlsx'
    wb.save(response)
    return response

export_messages_to_excel.short_description = "📤 Export selected messages to Excel"


# ✅ Sensor Data Admin
class SensorDataAdmin(admin.ModelAdmin):
    readonly_fields = [
        'timestamp',
        'T_cold', 'Humidity_cold',
        'T_hot_rack1', 'Humidity_rack1',
        'T_hot_rack2', 'Humidity_rack2',
        'T_hot_rack3', 'Humidity_rack3',
        'T_room', 'Room_Humidity',
        'P_total_room', 'P_total_cooling_system',
        'P_rack1', 'P_rack2'
    ]

    list_display = (
        'timestamp',
        'T_cold', 'Humidity_cold',
        'T_room', 'Room_Humidity',
        'P_total_room', 'P_total_cooling_system',
    )

    actions = [export_to_excel]
    change_list_template = "admin/sensor_data_changelist.html"

    def has_add_permission(self, request):
        return False

    def has_change_permission(self, request, obj=None):
        return False

    def has_delete_permission(self, request, obj=None):
        return False

    def export_filtered(self, request, period=None, start=None, end=None):
        qs = SensorData.objects.all()

        if period == "week":
            qs = qs.filter(timestamp__gte=now() - timedelta(days=7))
        elif period == "month":
            qs = qs.filter(timestamp__gte=now() - timedelta(days=30))
        elif period == "custom" and start and end:
            try:
                start_date = datetime.strptime(start, "%Y-%m-%d")
                end_date = datetime.strptime(end, "%Y-%m-%d")
                qs = qs.filter(timestamp__date__gte=start_date,
                               timestamp__date__lte=end_date)
            except ValueError:
                pass

        return export_to_excel(self, request, qs)

    def changelist_view(self, request, extra_context=None):
        if "export" in request.GET:
            return self.export_filtered(
                request,
                request.GET.get("export"),
                request.GET.get("start_date"),
                request.GET.get("end_date"),
            )
        elif "export_all" in request.GET:
            return export_to_excel(self, request, SensorData.objects.all())

        return super().changelist_view(request, extra_context)


# ✅ Custom User Admin
class CustomUserAdmin(UserAdmin):
    model = CustomUser
    list_display = ('email', 'is_staff', 'is_superuser', 'last_login', 'date_joined')
    list_filter = ('is_staff', 'is_superuser', 'is_active')
    search_fields = ('email',)
    ordering = ('email',)

    fieldsets = (
        (None, {"fields": ("email", "password")}),
        ("Permissions", {"fields": ("is_staff", "is_active", "user_permissions")}),
        ("Important dates", {"fields": ("last_login", "date_joined")}),
    )

    add_fieldsets = (
        (None, {
            "classes": ("wide",),
            "fields": ("email", "password1", "password2", "is_staff", "is_active")}
        ),
    )

    filter_horizontal = ("user_permissions",)

    def get_readonly_fields(self, request, obj=None):
        ro_fields = super().get_readonly_fields(request, obj)
        if obj:
            return ro_fields + ("is_superuser",)
        return ro_fields

    def has_delete_permission(self, request, obj=None):
        return request.user.is_superuser


# ✅ Login History Admin
class LoginHistoryAdmin(admin.ModelAdmin):
    list_display = ("user", "timestamp", "ip_address")
    list_filter = ("timestamp",)

    def has_add_permission(self, request):
        return False

    def has_change_permission(self, request, obj=None):
        return False

    def has_delete_permission(self, request, obj=None):
        return False


# ✅ Message Admin
class MessageAdmin(admin.ModelAdmin):
    list_display = ("timestamp", "type", "content", "source")
    list_filter = ("type", "timestamp", "source")
    search_fields = ("content", "source")
    # Remove default search bar
    search_fields = ()

    readonly_fields = ("timestamp", "type", "content", "data", "source")

        # Add Excel export action
    actions = [export_messages_to_excel]
        # ✅ Tell Django to use your custom template
    change_list_template = "admin/monitoring/message/change_list.html"

    def changelist_view(self, request, extra_context=None):
        if "export_all" in request.GET:
            return export_messages_to_excel(self, request, Message.objects.all())
        return super().changelist_view(request, extra_context)
    
    def has_add_permission(self, request):
        return False

    def has_change_permission(self, request, obj=None):
        return False

    def has_delete_permission(self, request, obj=None):
        return False


# --- Register models with custom site ---
custom_admin_site.register(SensorData, SensorDataAdmin)
custom_admin_site.register(CustomUser, CustomUserAdmin)
custom_admin_site.register(LoginHistory, LoginHistoryAdmin)
custom_admin_site.register(Message, MessageAdmin)
